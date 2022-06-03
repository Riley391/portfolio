import openpyxl
import pandas as pd
import os
import sys
from pathlib import Path
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter

def getFileTitle(inOrOut):
    if inOrOut == 'in':
        message = 'Please enter the name of the csv file to take data from: '
        fileEnding = '.csv'
    elif inOrOut == 'out':
        message = 'Please enter the name of your new Excel file here: '
        fileEnding = '.xlsx'
    resultFile = input(message)
    if not resultFile.endswith(fileEnding): resultFile += fileEnding
    return resultFile

def buildTrueOrFalse(string, inOrOut):
    if string == 'y':
        return True
    elif string == 'n':
        return False
    else:
        getFileTitle(inOrOut)

def createOutputFile(outputFile):
    if os.path.isfile(outputFile):
        deleteFile = input('This file already exists, would you like to overwrite it? (y/n): ')
        if deleteFile == 'y':
            os.remove(outputFile)
            print('File overwritten')
        elif deleteFile == 'n':
            outputFile = getFileTitle('out')
    Path(outputFile).touch()

def setBorder(ws):
    thin = Side(border_style='thin', color='000000')
    thinBorder = Border(top=thin, right=thin, bottom=thin, left=thin)
    for row in ws.iter_rows(1, ws.max_row):
        for cell in row:
            cell.border = thinBorder

def setColumnWidth(ws, width):
    column_widths = []
    for row in ws.iter_rows(1, ws.max_row):
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    column_widths[i] = width
            else:
                column_widths += [width]
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width

os.system('cls') if sys.platform.startswith('win') else os.system('clear')

inputFile = getFileTitle('in')
outputFile = getFileTitle('out')

createOutputFile(outputFile)

df = pd.read_csv(inputFile)
df.drop("Quiz Type", axis=1, inplace=True)
df.to_excel(outputFile, index=False)

wb = openpyxl.load_workbook(outputFile)
ws = wb['Sheet1']

setBorder(ws)
setColumnWidth(ws, 25)

wb.save(outputFile)